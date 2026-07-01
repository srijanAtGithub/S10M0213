"""
cowork_tools.py
--------------
Sandboxed filesystem tools for `sicily start`.

Mirrors the @modelcontextprotocol/server-filesystem interface,
re-implemented in pure Python with zero extra dependencies.

ALL tools are locked to a single root directory (the cwd where
`sicily start` was invoked). No path can escape that root.

Tool tiers
----------
Read-only tools  — safe:      read_file, list_directory, file_tree_shallow,
                              get_file_info, list_allowed_directories, read_file_lines
Write tools      — safe-ish: create_text_file, make_directory, edit_file_lines
                  Guarantee: never delete existing content.
                  edit_file_lines requires dry_run=False to apply changes.
Path pins        — memory:   pin_path, recall_path, recall_all_pins
                  Survive context summarisation — stored in process memory,
                  not in the message list.
"""

import datetime
import stat
from pathlib import Path
from typing import Optional
import importlib

from langchain_core.tools import tool

# Noise directories — skipped in trees and searches
SKIP_DIRS = {
    ".venv", "venv", "env", ".env",
    "node_modules",
    "__pycache__",
    ".git",
    ".mypy_cache", ".pytest_cache", ".ruff_cache",
    "dist", "build", ".eggs",
    ".tox", ".nox",
    ".idea", ".vscode",
    ".sicily-trash",
}


# Allowed extensions for text-based file reads and writes
# Binary formats (.docx, .xlsx, .pdf, …) are intentionally excluded from
# write operations — they require structured serialisation, not raw text I/O.
ALLOWED_WRITE_EXTENSIONS: frozenset[str] = frozenset({
    # Documents & notes
    ".txt", ".md", ".markdown", ".rst", ".org", ".tex",
    # Config & data interchange
    ".json", ".jsonl", ".ndjson",
    ".yaml", ".yml", ".toml",
    ".ini", ".cfg", ".conf", ".env",
    # Web & markup
    ".html", ".htm", ".css", ".scss", ".sass", ".xml", ".svg",
    # Source code — common languages
    ".py", ".pyi",
    ".js", ".mjs", ".cjs", ".ts", ".tsx", ".jsx",
    ".sh", ".bash", ".zsh", ".fish",
    ".rb", ".go", ".rs",
    ".java", ".kt", ".scala",
    ".c", ".cpp", ".cc", ".h", ".hpp",
    ".cs", ".fs",
    ".php", ".lua", ".r", ".sql",
    # Data & logs
    ".csv", ".tsv", ".log",
    # Misc text
    ".diff", ".patch", ".gitignore", ".editorconfig",
})


# Sandbox root
_SANDBOX_ROOT: Optional[Path] = None


def set_sandbox_root(path: Path) -> None:
    global _SANDBOX_ROOT
    _SANDBOX_ROOT = path.resolve()


def get_sandbox_root() -> Path:
    if _SANDBOX_ROOT is None:
        raise RuntimeError("Sandbox root has not been set. Call set_sandbox_root() first.")
    return _SANDBOX_ROOT


# Path pin store — survives context summarisation
# Stored in process memory, not in the message list, so the summariser
# cannot compress it away.
_PATH_PINS: dict[str, str] = {}


# Internal helpers
def _safe_path(relative: str) -> Path:
    """
    Resolve a user/AI-supplied path against the sandbox root.
    Raises PermissionError if the resolved path would escape the root.
    """
    root = get_sandbox_root()
    candidate = root / relative
    try:
        resolved = candidate.resolve()
    except OSError:
        # On Windows, resolve() can raise FileNotFoundError for paths
        # that don't exist yet. Fall back to normpath-based resolution,
        # which works for non-existent paths.
        import os
        resolved = Path(os.path.normpath(candidate))

    if not resolved.is_relative_to(root):
        raise PermissionError(
            f"Access denied: '{relative}' resolves outside the allowed directory."
        )
    return resolved


def _is_skipped(path: Path) -> bool:
    """True if this is a noise directory that should be excluded."""
    return path.is_dir() and path.name in SKIP_DIRS


def _fmt_ts(ts: float) -> str:
    return datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


def _fmt_permissions(mode: int) -> str:
    """Convert a stat st_mode integer to a human-readable 'rwxrwxrwx' string."""
    result = []
    for who in ("USR", "GRP", "OTH"):
        for perm, letter in (("R", "r"), ("W", "w"), ("X", "x")):
            flag = getattr(stat, f"S_I{perm}{who}")
            result.append(letter if mode & flag else "-")
    return "".join(result)


# Extensions that require binary parsing rather than UTF-8 text reads
_BINARY_EXTENSIONS = frozenset({".pdf", ".xlsx", ".xls", ".docx", ".doc"})


def _read_binary(path: Path) -> str:
    """
    Extract human-readable text from binary file formats.
    Dispatches to the appropriate parser based on file extension.
    Raises ImportError with an install hint if the required library is missing.
    Raises ValueError for unsupported binary extensions.
    """
    ext = path.suffix.lower()

    if ext == ".pdf":
        if importlib.util.find_spec("pdfplumber") is None:
            raise ImportError("pip install pdfplumber")
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
        return "\n\n".join(f"[Page {i+1}]\n{text}" for i, text in enumerate(pages) if text.strip())

    if ext in {".xlsx", ".xls"}:
        if importlib.util.find_spec("openpyxl") is None:
            raise ImportError("pip install openpyxl")
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [
                "\t".join("" if cell.value is None else str(cell.value) for cell in row)
                for row in ws.iter_rows()
            ]
            sheets.append(f"[Sheet: {name}]\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets)

    if ext in {".docx", ".doc"}:
        if importlib.util.find_spec("docx") is None:
            raise ImportError("pip install python-docx")
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    raise ValueError(
        f"No binary reader available for '{ext}'. "
        "For plain text files this tool reads UTF-8 directly. "
        "For other binary formats, a dedicated tool may be needed."
    )


# READ-ONLY TOOLS
@tool
def search_index(query: str) -> str:
    """
    Search the local RAG index for content relevant to a query.
 
    This is the FIRST tool to call for any question that involves finding
    information inside files — before reading any file directly.
 
    The index covers all text-based files in the sandbox:
    .txt, .md, .pdf, .docx, .xlsx, .py, .json, .csv, and more.
 
    Returns the top matching snippets with their file path and position.
    If a snippet looks relevant, use read_file_lines to read more context
    around it in the original file.
 
    Args:
        query: Plain-language description of what you are looking for.
               e.g. "quarterly budget figures" or "meeting notes from January"
    """
    from Cowork.cowork_rag import get_rag   # adjust import path to match your project
    rag = get_rag()
    if rag is None:
        return "RAG index is not initialised. This is a bug — please report it."
    results = rag.search(query)
    return rag.format_results(results)


@tool
def read_file(path: str, head: int = 0, tail: int = 0) -> str:
    """
    Read the contents of any file and return it as plain text.

    RECOMMENDED WORKFLOW:
    - Start with `head=50` to understand the file's structure and size.
    - When search_index (RAG) returns specific line numbers, prefer `read_file_lines`.
    - Use full read (`head=0, tail=0`) only for small-to-medium files or when you 
      genuinely need the entire content (e.g. small scripts, configs, short notes).

    Handles two categories transparently:

    Text-based files (.txt, .md, .py, .json, .csv, .yaml, .html, etc.)
        Raw UTF-8 content is returned as-is.

    Binary document formats
        .pdf   — text extracted page by page, labelled [Page N]
        .docx  — all paragraph text extracted in order
        .xlsx/.xls — every sheet as tab-separated table, labelled [Sheet: name]

    IMPORTANT:
    - Always start with `head=50` on unknown or potentially large files.
    - Avoid full reads on large files (logs, big CSVs, long source files, etc.).
      Use `read_file_lines` with targeted ranges or multiple calls instead.
    - Full reads are mainly justified for summarization of small files, 
      code review of scripts, or when the complete content is genuinely required.

    For precise line-range reading (especially after RAG), use `read_file_lines`.

    Args:
        path: Relative path to the file.
        head: If > 0, return only the first N lines.
        tail: If > 0, return only the last N lines.
              Cannot be combined with head.
    """
    if head > 0 and tail > 0:
        return "Error: Cannot specify both `head` and `tail` simultaneously."

    try:
        file_path = _safe_path(path)
    except PermissionError as e:
        return str(e)

    if not file_path.exists():
        return f"File '{path}' does not exist."
    if not file_path.is_file():
        return f"'{path}' is a directory, not a file."

    # Binary formats — route to dedicated parser
    if file_path.suffix.lower() in _BINARY_EXTENSIONS:
        try:
            content = _read_binary(file_path)
        except ImportError as e:
            return f"Cannot read '{path}': missing required package — {e}"
        except Exception as e:
            return f"Could not extract text from '{path}': {e}"

    # Text files — UTF-8 read
    else:
        try:
            content = file_path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            return f"Could not read '{path}': {e}"

    if head > 0:
        return "".join(content.splitlines(keepends=True)[:head])
    if tail > 0:
        return "".join(content.splitlines(keepends=True)[-tail:])

    return content


@tool
def read_file_lines(path: str, start_line: int, end_line: int) -> str:
    """
    Read a specific line range from a text file, with line numbers shown.
    Lines are 1-indexed and inclusive on both ends. Maximum 500 lines per call.

    Use this BEFORE edit_file_lines to confirm you are targeting the correct
    lines. This lets you work surgically on large files without pulling their
    full content into context.

    Typical workflow
    ----------------
    1. read_file(path, head=50)           — understand structure, find region
    2. read_file_lines(path, N, M)        — confirm exact lines before editing
    3. edit_file_lines(path, N, M, ...)   — make the surgical replacement

    Args:
        path:       Relative path to the file.
        start_line: First line to read (1-indexed).
        end_line:   Last line to read (inclusive).
    """
    if start_line < 1:
        return "Error: start_line must be >= 1."
    if end_line < start_line:
        return "Error: end_line must be >= start_line."
    if end_line - start_line > 500:
        return "Error: Cannot read more than 500 lines at once. Narrow your range."

    try:
        target = _safe_path(path)
    except PermissionError as e:
        return str(e)

    if not target.exists():
        return f"File '{path}' does not exist."
    if not target.is_file():
        return f"'{path}' is a directory, not a file."

    try:
        all_lines = target.read_text(encoding="utf-8", errors="replace").splitlines(keepends=True)
    except Exception as e:
        return f"Could not read '{path}': {e}"

    total = len(all_lines)
    if start_line > total:
        return f"File only has {total} lines. start_line={start_line} is out of range."

    actual_end = min(end_line, total)
    selected = all_lines[start_line - 1 : actual_end]

    numbered = "".join(f"{start_line + i:>6}  {line}" for i, line in enumerate(selected))
    header = f"[{path} | lines {start_line}–{actual_end} of {total}]\n"
    return header + numbered


@tool
def list_directory(path: str = ".") -> str:
    """
    List the immediate contents of a directory.
    Each entry is prefixed with [FILE] or [DIR].
    Does NOT recurse into subdirectories.

    Args:
        path: Relative path to the directory. Defaults to "." (sandbox root).
    """
    try:
        target = _safe_path(path)
    except PermissionError as e:
        return str(e)

    if not target.exists():
        return f"Directory '{path}' does not exist."
    if not target.is_dir():
        return f"'{path}' is a file, not a directory."

    try:
        entries = sorted(target.iterdir(), key=lambda p: (p.is_file(), p.name))
    except PermissionError:
        return f"Permission denied: cannot list '{path}'."

    if not entries:
        return "Directory is empty."

    lines = []
    for entry in entries:
        tag = "[DIR] " if entry.is_dir() else "[FILE]"
        note = "  [skipped — noise dir]" if _is_skipped(entry) else ""
        lines.append(f"{tag} {entry.name}{note}")

    return "\n".join(lines)


@tool
def file_tree_shallow(subdirectory: str = ".", max_depth: int = 2, max_entries: int = 200) -> str:
    """
    Show a recursive visual tree of files and folders — token-safe.

    Unlike an unlimited tree dump, this tool is depth-limited and entry-capped
    so it never floods the context window on large or deeply nested projects.

    Use this as the FIRST tool when exploring any unknown project. If you need
    to go deeper into a specific subdirectory, call again with that path as the
    root and a higher max_depth.

    Rules of thumb
    --------------
    - Start at root with max_depth=2 to understand the project shape.
    - Drill into a specific folder: file_tree_shallow("src/billing", max_depth=3)
    - When the cap warning appears, narrow the subdirectory instead of raising it.
    - Do NOT use this with max_depth > 4 on the project root.

    Args:
        subdirectory: Relative path to start the tree from. Defaults to ".".
        max_depth:    How many levels deep to recurse. Default 2. Hard max 6.
        max_entries:  Stop emitting after this many entries to prevent token
                      floods. Default 200. A warning is appended when hit.
    """
    try:
        target = _safe_path(subdirectory)
    except PermissionError as e:
        return str(e)

    if not target.exists():
        return f"Directory '{subdirectory}' does not exist."
    if not target.is_dir():
        return f"'{subdirectory}' is a file, not a directory."

    max_depth = min(max(1, max_depth), 6)
    root = get_sandbox_root()
    label = str(target.relative_to(root)) if subdirectory != "." else "."
    lines = [f"📁 {label}"]
    entry_count = 0
    capped = False

    def _render(directory: Path, prefix: str = "", depth: int = 0) -> None:
        nonlocal entry_count, capped
        if capped or depth >= max_depth:
            return
        try:
            children = sorted(directory.iterdir(), key=lambda p: (p.is_file(), p.name))
        except PermissionError:
            lines.append(f"{prefix}└── [permission denied]")
            return

        for i, child in enumerate(children):
            if capped:
                return
            is_last = i == len(children) - 1
            connector = "└── " if is_last else "├── "

            if _is_skipped(child):
                lines.append(f"{prefix}{connector}📁 {child.name}/  [skipped]")
                entry_count += 1
                if entry_count >= max_entries:
                    capped = True
                continue

            icon = "📁 " if child.is_dir() else "📄 "
            lines.append(f"{prefix}{connector}{icon}{child.name}")
            entry_count += 1
            if entry_count >= max_entries:
                capped = True
                return

            if child.is_dir():
                extension = "    " if is_last else "│   "
                _render(child, prefix + extension, depth + 1)

    _render(target)

    if capped:
        lines.append(
            f"\n⚠️  Output capped at {max_entries} entries. "
            "Narrow the subdirectory path to explore deeper sections."
        )

    lines.append(f"\n[Depth: {max_depth} | Entries shown: {entry_count}]")
    return "\n".join(lines)


@tool
def get_file_info(path: str) -> str:
    """
    Get detailed metadata about a file or directory.
    Returns: name, type, size, permissions, created, modified, and accessed times.

    Args:
        path: Relative path to the file or directory.
    """
    try:
        target = _safe_path(path)
    except PermissionError as e:
        return str(e)

    if not target.exists():
        return f"'{path}' does not exist."

    try:
        s = target.stat()
    except PermissionError:
        return f"Permission denied: cannot stat '{path}'."

    kind = "Directory" if target.is_dir() else "File"
    size = f"{s.st_size:,} bytes" if target.is_file() else "—"
    permissions = _fmt_permissions(s.st_mode)

    # Creation time:
    #   macOS  → st_birthtime (real creation time)
    #   Windows→ st_ctime     (real creation time)
    #   Linux  → st_ctime     (last metadata change; true birthtime not exposed by Python)
    created = _fmt_ts(getattr(s, "st_birthtime", s.st_ctime))

    return "\n".join([
        f"Name:        {target.name}",
        f"Type:        {kind}",
        f"Size:        {size}",
        f"Permissions: {permissions}",
        f"Created:     {created}",
        f"Modified:    {_fmt_ts(s.st_mtime)}",
        f"Accessed:    {_fmt_ts(s.st_atime)}",
        f"Path:        {path}",
    ])


@tool
def list_allowed_directories() -> str:
    """
    List all directories the agent is allowed to access.
    Returns the sandbox root that was locked in when `sicily start` was invoked.
    No input required.
    """
    root = get_sandbox_root()
    return f"Allowed directories:\n  {root}"


# PATH PIN TOOLS — survive context summarisation
@tool
def pin_path(alias: str, path: str) -> str:
    """
    Save a file path under a short alias so it survives context summarisation.

    The summariser compresses old tool outputs out of the message list. A path
    discovered 10 turns ago may no longer be in context when you need to act
    on it. Pinned paths live in process memory — the summariser cannot touch them.

    ALWAYS call this immediately after finding a file you plan to use later,
    before reading it or doing anything else.

    Examples
    --------
    pin_path("target",  "src/billing/formatters/pdf_renderer.py")
    pin_path("config",  "infrastructure/k8s/prod/values.yaml")
    pin_path("tests",   "tests/unit/billing/test_invoice.py")

    Args:
        alias: A short memorable name for this path (e.g. "target", "config").
        path:  The relative file path to save.
    """
    _PATH_PINS[alias] = path
    return f"📌 Pinned '{alias}' → '{path}'. Use recall_path('{alias}') to retrieve it later."


@tool
def recall_path(alias: str) -> str:
    """
    Retrieve a previously pinned file path by its alias.

    Use this whenever you need to act on a file but cannot be certain its
    path is still in your active context (it may have been summarised away).

    Args:
        alias: The alias used when pin_path was called.
    """
    if alias not in _PATH_PINS:
        all_pins = ", ".join(f"'{k}'" for k in _PATH_PINS) if _PATH_PINS else "none"
        return (
            f"No path pinned under alias '{alias}'. "
            f"Available pins: {all_pins}. "
            "If you have not pinned this path yet, use find_files_by_name to locate it first."
        )
    return f"📌 '{alias}' → '{_PATH_PINS[alias]}'"


@tool
def recall_all_pins() -> str:
    """
    List every currently pinned path.

    Call this at the start of any multi-step task to remind yourself what
    files you have already located, or after a long chain of tool calls
    to re-orient before taking a write action.
    """
    if not _PATH_PINS:
        return "No paths are currently pinned. Use pin_path to save file locations."
    lines = [f"  {alias:20s} → {path}" for alias, path in _PATH_PINS.items()]
    return "📌 Pinned paths:\n" + "\n".join(lines)


# WRITE TOOLS
@tool
def create_text_file(
    path: str,
    content: str,
    create_parents: bool = True,
) -> str:
    """
    Create a NEW text file at the given relative path with the provided content.

    Safety guarantees
    -----------------
    - Will NEVER overwrite an existing file or directory. If the path already
      exists the operation is aborted immediately and an error is returned.
    - The resolved path must stay inside the sandbox root; any traversal attempt
      (e.g. "../../etc/passwd") is blocked before any I/O occurs.
    - Only recognised text-based extensions are accepted (see list below).
    - Parent directories are created automatically when `create_parents=True`
      (the default), as long as they remain inside the sandbox.

    To edit an existing file, use edit_file_lines instead.

    Supported extensions
    --------------------
    Documents/notes : .txt .md .markdown .rst .org .tex
    Config/data     : .json .jsonl .ndjson .yaml .yml .toml .ini .cfg .conf .env
    Web/markup      : .html .htm .css .scss .sass .xml .svg
    Source code     : .py .pyi .js .mjs .cjs .ts .tsx .jsx .sh .bash .zsh .fish
                      .rb .go .rs .java .kt .scala .c .cpp .cc .h .hpp .cs .fs
                      .php .lua .r .sql
    Data/logs       : .csv .tsv .log
    Misc text       : .diff .patch .gitignore .editorconfig

    Args:
        path:           Relative path for the new file, including its name and
                        extension (e.g. "notes/meeting.md").
        content:        UTF-8 text content to write.
        create_parents: When True (default), any missing parent directories are
                        created automatically. Set to False if you want the
                        operation to fail when a parent does not exist.
    """
    # 1. Sandbox enforcement
    try:
        target = _safe_path(path)
    except PermissionError as e:
        return str(e)

    # 2. No-overwrite guard
    if target.exists():
        kind = "directory" if target.is_dir() else "file"
        return (
            f"Refused: '{path}' already exists as a {kind}. "
            "Use edit_file_lines to modify an existing file."
        )

    # 3. Extension whitelist
    ext = target.suffix.lower()
    if not ext:
        return (
            f"Refused: '{path}' has no file extension. "
            "Please include one (e.g. report.md, config.yaml)."
        )
    if ext not in ALLOWED_WRITE_EXTENSIONS:
        allowed_str = "  " + "\n  ".join(sorted(ALLOWED_WRITE_EXTENSIONS))
        return (
            f"Refused: extension '{ext}' is not in the allowed list.\n"
            f"Supported extensions:\n{allowed_str}"
        )

    # 4. Parent directory handling
    parent = target.parent
    if not parent.exists():
        if not create_parents:
            rel_parent = parent.relative_to(get_sandbox_root())
            return (
                f"Error: parent directory '{rel_parent}' does not exist. "
                "Pass create_parents=True to create it automatically, "
                "or use make_directory first."
            )
        try:
            parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return f"Could not create parent directories for '{path}': {e}"

    # 5. Write
    try:
        target.write_text(content, encoding="utf-8")
    except Exception as e:
        return f"Could not write '{path}': {e}"

    size = target.stat().st_size
    return (
        f"Created '{path}'.\n"
        f"Size: {size:,} bytes | Encoding: utf-8"
    )


@tool
def edit_file_lines(
    path: str,
    start_line: int,
    end_line: int,
    new_content: str,
    dry_run: bool = True,
) -> str:
    """
    Replace a specific line range in an existing file with new content.
    This is the correct tool for any task that modifies an existing file.

    Safety design
    -------------
    - dry_run=True (the default): shows a diff-style preview WITHOUT writing.
      Always call with dry_run=True first so the user can confirm the change.
    - dry_run=False: actually writes the change. Only use after the user
      confirms the preview is correct.
    - The file MUST already exist. Use create_text_file for new files.
    - Replaces lines [start_line, end_line] inclusive (1-indexed) with
      new_content. All surrounding lines are untouched.
    - Set new_content="" to delete the target lines without inserting anything.
    - Only text-based extensions (same list as create_text_file) are supported.

    Typical workflow
    ----------------
    1. read_file(path, head=50)                         — understand structure
    2. read_file_lines(path, N, M)                      — confirm exact target
    3. edit_file_lines(path, N, M, new, dry_run=True)   — preview the change
    4. User confirms the preview looks correct
    5. edit_file_lines(path, N, M, new, dry_run=False)  — apply it

    Args:
        path:        Relative path to the file.
        start_line:  First line to replace (1-indexed).
        end_line:    Last line to replace (inclusive).
        new_content: Replacement text. Pass "" to delete the range entirely.
        dry_run:     If True (default), show a preview without writing anything.
    """
    if start_line < 1:
        return "Error: start_line must be >= 1."
    if end_line < start_line:
        return "Error: end_line must be >= start_line."

    try:
        target = _safe_path(path)
    except PermissionError as e:
        return str(e)

    if not target.exists():
        return (
            f"File '{path}' does not exist. "
            "Use create_text_file to create new files."
        )
    if not target.is_file():
        return f"'{path}' is a directory, not a file."

    # Extension guard — only edit text-based files
    ext = target.suffix.lower()
    if ext not in ALLOWED_WRITE_EXTENSIONS:
        return (
            f"Refused: extension '{ext}' is not in the allowed list for editing. "
            "Only text-based files can be edited."
        )

    try:
        all_lines = target.read_text(encoding="utf-8", errors="replace").splitlines(keepends=True)
    except Exception as e:
        return f"Could not read '{path}': {e}"

    total = len(all_lines)
    if start_line > total + 1:
        return f"File only has {total} lines. start_line={start_line} is out of range."

    actual_end = min(end_line, total)
    removed = all_lines[start_line - 1 : actual_end]

    # Ensure new_content ends with a newline so the file stays well-formed
    if new_content and not new_content.endswith("\n"):
        replacement_block = new_content + "\n"
    else:
        replacement_block = new_content

    new_file_lines = all_lines[: start_line - 1] + ([replacement_block] if replacement_block else []) + all_lines[actual_end:]
    new_content_full = "".join(new_file_lines)

    if dry_run:
        removed_preview = (
            "".join(f"  - {l.rstrip()}\n" for l in removed)
            or "  (nothing — pure insertion before this line)\n"
        )
        added_lines = replacement_block.splitlines() if replacement_block else []
        added_preview = (
            "".join(f"  + {l}\n" for l in added_lines)
            if added_lines
            else "  (deleted — no replacement)\n"
        )
        return (
            f"[DRY RUN — no changes written]\n\n"
            f"File:   {path}\n"
            f"Range:  lines {start_line}–{actual_end} of {total}\n\n"
            f"REMOVE:\n{removed_preview}\n"
            f"INSERT:\n{added_preview}\n"
            f"Call again with dry_run=False to apply."
        )

    # Apply the edit
    try:
        target.write_text(new_content_full, encoding="utf-8")
    except Exception as e:
        return f"Could not write '{path}': {e}"

    new_total = len(new_file_lines)
    delta = new_total - total
    delta_str = f"+{delta}" if delta >= 0 else str(delta)
    added_count = len(replacement_block.splitlines()) if replacement_block else 0
    return (
        f"✅ Edit applied to '{path}'.\n"
        f"Replaced lines {start_line}–{actual_end} "
        f"({len(removed)} line(s) removed → {added_count} line(s) inserted).\n"
        f"File now has {new_total} lines ({delta_str})."
    )


@tool
def make_directory(path: str) -> str:
    """
    Create a new directory at the given relative path, including any missing
    intermediate parents. Idempotent: succeeds silently if the directory
    already exists.

    Safety guarantees
    -----------------
    - Will NOT fail or overwrite if the directory already exists.
    - Will NOT touch any existing files or directories inside the path.
    - The resolved path must stay inside the sandbox root.
    - Will refuse if the path already exists as a *file*.

    Args:
        path: Relative path of the directory to create (e.g. "reports/q3").
    """
    # ── 1. Sandbox enforcement ────────────────────────────────────────────────
    try:
        target = _safe_path(path)
    except PermissionError as e:
        return str(e)

    # ── 2. Collision check ────────────────────────────────────────────────────
    if target.is_file():
        return (
            f"Refused: '{path}' already exists as a file. "
            "Cannot create a directory at that path."
        )

    if target.is_dir():
        return f"Directory '{path}' already exists — nothing to do."

    # ── 3. Create ─────────────────────────────────────────────────────────────
    try:
        target.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return f"Could not create directory '{path}': {e}"

    return f"Directory '{path}' created."


# EXPORTED TOOL LIST
LOCAL_TOOLS = [
    # Read-only (safe)
    search_index,
    read_file,
    read_file_lines,
    list_directory,
    file_tree_shallow,
    get_file_info,
    list_allowed_directories,

    # Path pins (process memory — survive summarisation)
    pin_path,
    recall_path,
    recall_all_pins,

    # Write (safe-ish)
    create_text_file,
    edit_file_lines,
    make_directory,
]


# SPINNER STATUS MESSAGES
TOOL_STATUS_MAP = {
    "search_index": lambda args: (
        f"Searching index for [white]'{args.get('query')}'[/white]"
    ),
    "read_file": lambda args: (
        f"Reading first {args.get('head')} lines of [white]'{args.get('path')}'[/white]"
        if args.get("head")
        else f"Reading file [white]'{args.get('path')}'[/white]"
    ),
    "read_file_lines": lambda args: (
        f"Reading lines {args.get('start_line')}–{args.get('end_line')} of "
        f"[white]'{args.get('path')}'[/white]"
    ),
    "list_directory": lambda args: (
        f"Listing contents of [white]'{args.get('path', '.')}'[/white]"
    ),
    "file_tree_shallow": lambda args: (
        f"Scanning directory tree of [white]'{args.get('subdirectory', '.')}'[/white] "
        f"(depth {args.get('max_depth', 2)})"
    ),
    "get_file_info": lambda args: (
        f"Inspecting metadata for [white]'{args.get('path')}'[/white]"
    ),
    "list_allowed_directories": lambda args: "Checking sandbox boundary",
    "pin_path": lambda args: (
        f"Pinning [white]'{args.get('path')}'[/white] as [white]'{args.get('alias')}'[/white]"
    ),
    "recall_path": lambda args: (
        f"Recalling pinned path [white]'{args.get('alias')}'[/white]"
    ),
    "recall_all_pins": lambda args: "Checking all pinned paths",
    "create_text_file": lambda args: (
        f"Creating [white]'{args.get('path')}'[/white]"
    ),
    "edit_file_lines": lambda args: (
        f"Previewing edit to [white]'{args.get('path')}'[/white] "
        f"(lines {args.get('start_line')}–{args.get('end_line')})"
        if args.get("dry_run", True)
        else f"Applying edit to [white]'{args.get('path')}'[/white] "
             f"(lines {args.get('start_line')}–{args.get('end_line')})"
    ),
    "make_directory": lambda args: (
        f"Creating directory [white]'{args.get('path')}'[/white]"
    ),
}

import Cowork.cowork_tool_fileops as fileops
LOCAL_TOOLS.extend(fileops.FILEOPS_TOOLS)   # Merge fileops tools
TOOL_STATUS_MAP.update(fileops.FILEOPS_TOOL_STATUS_MAP) # Merge status messages


def get_friendly_tool_message(tool_call: dict) -> str:
    """Extracts the tool name and args to build a readable status update."""
    name = tool_call.get("name")
    args = tool_call.get("args", {})

    if name in TOOL_STATUS_MAP:
        return TOOL_STATUS_MAP[name](args)

    # Fallback for any future tools not yet in the map
    return name or "working..."
