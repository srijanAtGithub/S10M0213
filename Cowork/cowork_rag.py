"""
cowork_rag.py
-------------
Local RAG engine for Sicily Cowork.

Hybrid search architecture
--------------------------
  Stage 1 — TF-IDF (sklearn)
      Keyword pre-filter. Fully local, zero cost, zero latency.
      Scans all indexed chunks and returns top-N keyword matches.

  Stage 2 — ChromaDB (text-embedding-3-small)
      Semantic vector search. Understands meaning, not just words.
      Runs against the same corpus and returns top-N semantic matches.

  Stage 3 — Reciprocal Rank Fusion (RRF)
      Merges both ranked lists into a single, de-duplicated ranking.
      Neither layer alone wins — both inform the final order.

Storage layout (global, shared across all sessions)
--------------------------------------------------
  ~/.sicily/file-index/
      chroma/           ChromaDB persistent collection (all sessions share this)
      tfidf.pkl         Fitted TfidfVectorizer + sparse matrix + ID list
      registry.json     { "/abs/path/to/file.pdf": "2024-01-15T10:30:00" }
                          ↑ keyed by absolute path, not relative

Incremental updates
-------------------
  On every session start, Sicily:
    1. Runs a global cleanup — removes registry entries for files that no
       longer exist on disk, regardless of which session indexed them.
    2. Removes entries for files that were under the current sandbox but
       have since been moved or deleted.
    3. Walks the current sandbox and re-indexes only new or changed files.
  Files indexed by other sessions are never touched unless they are deleted
  from disk. Starting from a parent or child of a previously indexed
  directory reuses existing chunks — nothing is re-indexed unnecessarily.
"""

import hashlib
import json
import os
import pickle
import time
from pathlib import Path
from typing import Optional, Callable

import structlog
from langchain_chroma import Chroma
from langchain_core.documents import Document
from langchain_openai import OpenAIEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

log = structlog.get_logger()


# Constants
CHUNK_SIZE        = 500   # characters per chunk
CHUNK_OVERLAP     = 50    # overlap between adjacent chunks
TOP_K             = 5     # final results returned to the LLM
TFIDF_CANDIDATES  = 50    # how many TF-IDF hits to feed into RRF
SEMANTIC_CANDIDATES = 50  # how many ChromaDB hits to feed into RRF
RRF_K             = 60    # RRF constant (standard value, do not change lightly)

# Chroma's `where` clause supports exact-match / $in, but not string
# prefix matching — there's no native "give me everything under this
# folder" filter. So sandbox-scoped queries are done via `$in` against
# the exact file list, sent in batches of this size, rather than either
# (a) one unbounded global fetch, or (b) one single $in list that could
# grow into the thousands for a large sandbox.
CHROMA_QUERY_BATCH_SIZE = 200

# File types that get RAG-indexed (embedded + TF-IDF'd).
#
# Sicily Cowork's RAG layer is scoped to general / non-coding files only.
# Source code, markup, and project-config extensions are deliberately
# excluded here to avoid silently embedding entire codebases (cost +
# noise) when `sicily start` happens to be run inside a coding project.
# Code files are still fully readable/writable via the direct file tools
# (read_file, read_file_lines, create_text_file, ...) in cowork_tools.py —
# they are just never pushed into the vector/keyword index.
#
# A dedicated coding-project variant of Sicily Cowork (closer to a
# `claude cowork`-style tool) is planned separately to handle that case
# with its own indexing strategy.
INDEXABLE_EXTENSIONS: frozenset[str] = frozenset({
    # Plain documents / notes
    ".txt", ".md", ".markdown", ".rst", ".org", ".tex",
    # General data / logs (not project source)
    ".csv", ".tsv", ".log",
    # Binary formats with text extractors
    ".pdf", ".docx", ".doc", ".xlsx", ".xls",
})

# Explicitly NOT indexed (kept here as documentation, not used directly):
#   Config / data interchange : .json .jsonl .ndjson .yaml .yml .toml .ini .cfg .conf .env
#   Web / markup               : .html .htm .xml .css .scss
#   Source code                : .py .pyi .js .mjs .ts .tsx .jsx .sh .bash .zsh
#                                 .rb .go .rs .java .kt .scala .c .cpp .h .hpp
#                                 .cs .php .lua .r .sql

# Directories to skip when walking the sandbox
SKIP_DIRS: set[str] = {
    ".venv", "venv", "env", ".env",
    "node_modules", "__pycache__", ".git",
    ".mypy_cache", ".pytest_cache", ".ruff_cache",
    "dist", "build", ".eggs",
    ".tox", ".nox", ".idea", ".vscode",
    ".sicily-trash",
}


# Module-level singleton
# Set once in cowork_session.py, read by the search_index tool in cowork_tools.py
_RAG_INSTANCE: Optional["SicilyRAG"] = None


def set_rag(instance: "SicilyRAG") -> None:
    global _RAG_INSTANCE
    _RAG_INSTANCE = instance


def get_rag() -> Optional["SicilyRAG"]:
    return _RAG_INSTANCE


# Path helpers
def _global_rag_dir() -> Path:
    """
    Return (and create) the single global RAG storage directory.
    All sessions share this — the index is keyed by absolute file path,
    so starting from any directory reuses previously indexed files.
    """
    path = Path.home() / ".sicily" / "file-index"
    path.mkdir(parents=True, exist_ok=True)
    return path


# Text extraction
def _extract_text(file_path: Path) -> Optional[str]:
    """
    Extract plain text from any supported file type.
    Mirrors the extraction logic already in cowork_tools._read_binary,
    kept here to avoid a circular import.
    Returns None on failure or unsupported type.
    """
    ext = file_path.suffix.lower()
    if ext not in INDEXABLE_EXTENSIONS:
        return None

    try:
        # Binary formats
        if ext == ".pdf":
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                pages = [page.extract_text() or "" for page in pdf.pages]
            return "\n\n".join(
                f"[Page {i + 1}]\n{t}" for i, t in enumerate(pages) if t.strip()
            )

        if ext in {".xlsx", ".xls"}:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                rows = [
                    "\t".join(
                        "" if cell.value is None else str(cell.value)
                        for cell in row
                    )
                    for row in ws.iter_rows()
                ]
                sheets.append(f"[Sheet: {name}]\n" + "\n".join(rows))
            wb.close()
            return "\n\n".join(sheets)

        if ext in {".docx", ".doc"}:
            from docx import Document
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

        # Plain text (all remaining supported extensions)
        return file_path.read_text(encoding="utf-8", errors="replace")

    except Exception as exc:
        log.warning("rag.extract_failed", path=str(file_path), error=str(exc))
        return None


# Core RAG class
class SicilyRAG:
    """
    Local hybrid RAG engine for a single Sicily session.

    Lifecycle
    ---------
    1. Instantiate with the sandbox root.
    2. Call index_session() once at startup (blocks until done).
    3. Call search(query) for every user question that needs file content.
    """

    def __init__(self, sandbox_root: Path) -> None:
        self.sandbox_root = sandbox_root.resolve()
        self.rag_dir      = _global_rag_dir()

        # Paths
        self._chroma_dir    = self.rag_dir / "chroma"
        self._tfidf_path    = self.rag_dir / "tfidf.pkl"
        self._registry_path = self.rag_dir / "registry.json"

        # LangChain / ChromaDB
        self._embeddings = OpenAIEmbeddings(model="text-embedding-3-small")
        self._vectorstore = Chroma(
            collection_name="sicily_rag",
            embedding_function=self._embeddings,
            persist_directory=str(self._chroma_dir),
            collection_metadata={"hnsw:space": "cosine"},
        )
        self._splitter = RecursiveCharacterTextSplitter(
            chunk_size=CHUNK_SIZE,
            chunk_overlap=CHUNK_OVERLAP,
            length_function=len,
            add_start_index=True,   # stored in metadata as "start_index"
        )

        # Exact set of absolute paths that belong to THIS sandbox session.
        # Populated by index_session(). Used as the single source of truth
        # for scoping both TF-IDF and semantic search to the current sandbox —
        # set-membership, not string prefix matching (a prefix check like
        # `abs_path.startswith(sandbox_root)` is unsafe: "/proj" would wrongly
        # match a sibling directory "/proj2").
        self._sandbox_abs_paths: set[str] = set()

        # TF-IDF state
        self._tfidf_vectorizer: Optional[TfidfVectorizer] = None
        self._tfidf_matrix   = None
        self._tfidf_ids: list[str] = []          # row i  →  ChromaDB chunk id
        self._tfidf_id_to_abspath: dict[str, str] = {}  # chunk id → abs file path

        # File registry
        # { "/absolute/path/to/file.pdf": "2024-01-15T10:30:00" }
        self._registry: dict[str, str] = self._load_registry()


    # Registry
    def _load_registry(self) -> dict[str, str]:
        if self._registry_path.exists():
            try:
                return json.loads(self._registry_path.read_text(encoding="utf-8"))
            except Exception:
                return {}
        return {}


    def _save_registry(self) -> None:
        self._registry_path.write_text(
            json.dumps(self._registry, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )


    # TF-IDF persistence
    def _save_tfidf(self) -> None:
        with open(self._tfidf_path, "wb") as fh:
            pickle.dump(
                {
                    "vectorizer":      self._tfidf_vectorizer,
                    "matrix":          self._tfidf_matrix,
                    "ids":             self._tfidf_ids,
                    "id_to_abspath":   self._tfidf_id_to_abspath,
                },
                fh,
            )


    def _load_tfidf(self) -> bool:
        if not self._tfidf_path.exists():
            return False
        try:
            with open(self._tfidf_path, "rb") as fh:
                data = pickle.load(fh)
            self._tfidf_vectorizer     = data["vectorizer"]
            self._tfidf_matrix         = data["matrix"]
            self._tfidf_ids            = data["ids"]
            self._tfidf_id_to_abspath  = data.get("id_to_abspath", {})
            return True
        except Exception:
            return False


    def _fetch_chunks_for_paths(self, paths: list[str]) -> tuple[list[str], list[str], list[dict]]:
        """
        Fetch chunks from ChromaDB whose abs_path is in `paths`, querying
        in batches of CHROMA_QUERY_BATCH_SIZE instead of either (a) one
        unfiltered global .get() across every sandbox ever indexed, or
        (b) one single $in list that grows unbounded with sandbox size.

        Each batch query is filtered server-side via Chroma's `where`
        clause — this is the actual cost fix. A Python-side `if` check
        after a global fetch doesn't reduce the database round-trip;
        only filtering inside the query itself does that.

        Returns (ids, documents, metadatas), all three lists aligned
        by index, same shape as ChromaDB's raw .get() result.
        """
        ids, docs, metadatas = [], [], []
        if not paths:
            return ids, docs, metadatas

        for i in range(0, len(paths), CHROMA_QUERY_BATCH_SIZE):
            batch = paths[i : i + CHROMA_QUERY_BATCH_SIZE]
            result = self._vectorstore.get(
                where={"abs_path": {"$in": batch}},
                include=["documents", "metadatas"],
            )
            ids.extend(result.get("ids") or [])
            docs.extend(result.get("documents") or [])
            metadatas.extend(result.get("metadatas") or [])

        return ids, docs, metadatas


    def _rebuild_tfidf(self) -> None:
        """
        Rebuild the TF-IDF index from scratch, scoped to the CURRENT
        sandbox only — not the global cross-session corpus.

        ChromaDB itself stores chunks from every sandbox Sicily has ever
        indexed (that's intentional, for cross-session reuse — see module
        docstring). Fitting TF-IDF over that entire history would be
        wrong on two counts: it gets slower with every unrelated project
        you've ever opened, and term statistics from unrelated projects
        pollute relevance scoring for the project you're actually in.

        Chroma's `where` clause only supports exact-match / $in, not a
        native "starts with this folder" filter — so there's no single
        query that says "everything under /abc/xyz". Instead, the exact
        (already directory-walk-correct) file list in
        self._sandbox_abs_paths is sent to Chroma via _fetch_chunks_for_paths,
        which batches the $in filter so the query stays server-side-filtered
        and bounded, never an unfiltered global pull. This must run AFTER
        self._sandbox_abs_paths has been populated by index_session().

        Called unconditionally at the end of index_session() — see
        comment there for why "unconditional" is correct here (it's
        local/sklearn fitting, no embedding API cost, and now bounded
        by sandbox size rather than global corpus size).
        """
        ids, docs, metadatas = self._fetch_chunks_for_paths(
            list(self._sandbox_abs_paths)
        )

        if not docs:
            # Nothing indexed for this sandbox — clear TF-IDF state
            self._tfidf_vectorizer    = None
            self._tfidf_matrix        = None
            self._tfidf_ids           = []
            self._tfidf_id_to_abspath = {}
            return

        # Build ID → abs_path mapping (sandbox-scoped already, but kept
        # as a fast lookup for _tfidf_search's defensive filter)
        id_to_abspath = {
            cid: m.get("abs_path", "")
            for cid, m in zip(ids, metadatas)
        }

        vectorizer = TfidfVectorizer(
            strip_accents="unicode",
            analyzer="word",
            ngram_range=(1, 2),     # unigrams + bigrams (e.g. "Q3 budget")
            min_df=1,
            sublinear_tf=True,      # log(1+tf) dampens very frequent terms
        )
        matrix = vectorizer.fit_transform(docs)

        self._tfidf_vectorizer    = vectorizer
        self._tfidf_matrix        = matrix
        self._tfidf_ids           = list(ids)
        self._tfidf_id_to_abspath = id_to_abspath
        self._save_tfidf()

    # Chunk ID
    @staticmethod
    def _chunk_id(abs_path: str, chunk_index: int) -> str:
        """
        Stable, deterministic ID for a chunk.
        Used as the ChromaDB document ID so re-indexing is idempotent.
        Keyed by absolute path so the same file is never indexed twice
        regardless of which sandbox session discovered it.
        """
        raw = f"{abs_path}::{chunk_index}"
        return hashlib.md5(raw.encode()).hexdigest()


    # File-level indexing
    def _mtime_iso(self, file_path: Path) -> str:
        ts = file_path.stat().st_mtime
        return time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(ts))


    def _should_reindex(self, file_path: Path) -> bool:
        abs_path = str(file_path)
        if abs_path not in self._registry:
            return True                                           # new file
        return self._mtime_iso(file_path) != self._registry[abs_path]  # modified


    def _delete_file_chunks(self, abs_path: str) -> None:
        """Remove all ChromaDB documents that belong to one file."""
        try:
            existing = self._vectorstore.get(
                where={"abs_path": abs_path},
                include=[],
            )
            if existing["ids"]:
                self._vectorstore.delete(ids=existing["ids"])
        except Exception as exc:
            log.warning("rag.delete_chunks_failed", path=abs_path, error=str(exc))


    def _index_file(self, file_path: Path) -> int:
        """
        Index one file:
          1. Extract text.
          2. Delete any previously indexed chunks for this file.
          3. Split into chunks with character offsets.
          4. Convert character offsets → line numbers (1-based).
          5. Store in ChromaDB with full metadata (keyed by abs_path).
          6. Update registry.

        Returns number of chunks stored (0 if skipped).
        """
        abs_path = str(file_path)

        text = _extract_text(file_path)
        if not text or not text.strip():
            return 0

        # Remove stale chunks
        self._delete_file_chunks(abs_path)

        # Use create_documents to get metadata with start_index
        raw_docs = self._splitter.create_documents([text])
        if not raw_docs:
            return 0

        # ── Build line offset map for fast char → line conversion ──
        # line_starts[i] = character index where line (i+1) begins
        line_starts = [0]
        for i, ch in enumerate(text):
            if ch == "\n":
                line_starts.append(i + 1)

        def char_to_line(char_index: int) -> int:
            """Convert character offset to 1-based line number."""
            if not line_starts:
                return 1
            lo, hi = 0, len(line_starts) - 1
            while lo < hi:
                mid = (lo + hi + 1) // 2
                if line_starts[mid] <= char_index:
                    lo = mid
                else:
                    hi = mid - 1
            return lo + 1  # 1-indexed

        # ── Prepare chunks with line numbers ──
        mtime_iso = self._mtime_iso(file_path)
        n = len(raw_docs)

        texts = []
        ids = []
        metadatas = []

        for i, doc in enumerate(raw_docs):
            start_char = doc.metadata.get("start_index", 0)
            end_char = start_char + len(doc.page_content)

            start_line = char_to_line(start_char)
            end_line = char_to_line(end_char)

            texts.append(doc.page_content)
            ids.append(self._chunk_id(abs_path, i))
            metadatas.append({
                "abs_path":     abs_path,           # for filtering / deletion
                "file_name":    file_path.name,
                "extension":    file_path.suffix.lower(),
                "chunk_index":  i,
                "total_chunks": n,
                "last_modified": mtime_iso,
                "start_line":   start_line,
                "end_line":     end_line,
            })

        # Embed + store
        self._vectorstore.add_texts(
            texts=texts,
            metadatas=metadatas,
            ids=ids,
        )

        self._registry[abs_path] = mtime_iso
        return n


    # Session-level indexing
    def _walk_sandbox(self) -> list[Path]:
        """Return every indexable file inside the sandbox, skipping noise dirs."""
        files: list[Path] = []
        for dirpath, dirnames, filenames in os.walk(self.sandbox_root):
            dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS]
            for fname in filenames:
                fp = Path(dirpath) / fname
                if fp.suffix.lower() in INDEXABLE_EXTENSIONS:
                    files.append(fp)
        return files


    def _clean_stale_entries(self) -> int:
        """
        Global cleanup: remove registry entries for files that no longer
        exist on disk, regardless of which session originally indexed them.
        Run once at startup before any other indexing logic.
        """
        stale = [p for p in list(self._registry) if not Path(p).exists()]
        for abs_path in stale:
            self._delete_file_chunks(abs_path)
            del self._registry[abs_path]
        if stale:
            log.info("rag.stale_cleaned", count=len(stale))
        return len(stale)


    def count_pending(self) -> int:
        """
        Cheap upfront check: how many files WOULD be (re)indexed if
        index_session() ran right now — without doing any extraction or
        embedding work. Just a directory walk + mtime comparison against
        the registry.

        Used by callers to decide, before rendering anything, whether to
        show a progress bar (real work ahead) or a lightweight "up to
        date" message (nothing but cleanup/TF-IDF resync to do).
        """
        all_files = self._walk_sandbox()
        return sum(1 for fp in all_files if self._should_reindex(fp))


    def index_session(self, progress_callback: Optional[Callable[[Path, int, int], None]] = None,) -> dict:
        """
        Run at Sicily startup.  Walks the sandbox and brings the index
        in sync with the current state of the filesystem.

        Parameters
        ----------
        progress_callback   Optional. Called as
                            progress_callback(file_path, current_index, total_to_index)
                            right before each file that needs (re)indexing is
                            processed. total_to_index is fixed upfront, so
                            callers can render a real (determinate) progress
                            bar instead of an indefinite spinner.

        Returns
        -------
        dict with keys: total_files, indexed, skipped, deleted, failed
        """
        all_files   = self._walk_sandbox()
        current_abs = {str(f) for f in all_files}

        # Populate the single source of truth for "what belongs to this
        # sandbox" BEFORE any indexing/filtering logic runs below — both
        # _rebuild_tfidf() and search()'s semantic stage depend on this.
        self._sandbox_abs_paths = current_abs

        # Step 1 — global cleanup: files deleted from disk anywhere
        self._clean_stale_entries()

        # Step 2 — sandbox cleanup: files that were under this sandbox
        # but have since been moved or removed (still exist on disk elsewhere)
        deleted_paths = [
            p for p in list(self._registry)
            if Path(p).is_relative_to(self.sandbox_root) and p not in current_abs
        ]
        for abs_path in deleted_paths:
            self._delete_file_chunks(abs_path)
            del self._registry[abs_path]

        # Step 3 — index new / modified files.
        # Figure out the exact worklist UP FRONT so we have a real total
        # for a progress bar, instead of discovering the count as we go.
        to_index = [fp for fp in all_files if self._should_reindex(fp)]
        total_to_index = len(to_index)
        skipped = len(all_files) - total_to_index
        indexed = failed = 0

        for i, fp in enumerate(to_index, start=1):
            if progress_callback is not None:
                progress_callback(fp, i, total_to_index)
            try:
                n = self._index_file(fp)
                if n > 0:
                    indexed += 1
                    # debug, not info — with a progress bar (Rich Live)
                    # rendering, a stray per-file stdout write corrupts the
                    # display. One summary line at the end is enough.
                    # log.debug("rag.indexed", path=str(fp), chunks=n)
            except Exception as exc:
                failed += 1
                log.warning("rag.index_failed", path=str(fp), error=str(exc))

        self._save_registry()
        self._rebuild_tfidf()

        return {
            "total_files": len(all_files),
            "indexed":     indexed,
            "skipped":     skipped,
            "deleted":     len(deleted_paths),
            "failed":      failed,
        }


    # Search
    def _tfidf_search(self, query: str, k: int) -> list[tuple[str, float]]:
        """
        Keyword search. The TF-IDF matrix is already scoped to the current
        sandbox (see _rebuild_tfidf), so this filter is a defensive
        no-op in the normal case — kept as a second line of defense in
        case the matrix is ever stale. Uses exact set membership against
        self._sandbox_abs_paths, NOT string prefix matching: a prefix
        check like abs_path.startswith(sandbox_root) would incorrectly
        match a sibling directory (e.g. "/proj" matching "/proj2/...").
        Returns [(chunk_id, score), ...].
        """
        if self._tfidf_vectorizer is None or self._tfidf_matrix is None:
            return []
        try:
            q_vec  = self._tfidf_vectorizer.transform([query])
            scores = cosine_similarity(q_vec, self._tfidf_matrix).flatten()

            # Score all chunks, then filter to current sandbox only
            results = [
                (self._tfidf_ids[i], float(scores[i]))
                for i in range(len(self._tfidf_ids))
                if scores[i] > 0.0
                and self._tfidf_id_to_abspath.get(self._tfidf_ids[i], "") in self._sandbox_abs_paths
            ]
            results.sort(key=lambda x: x[1], reverse=True)
            return results[:k]

        except Exception as exc:
            log.warning("rag.tfidf_search_failed", error=str(exc))
            return []


    def _semantic_search(self, query: str, k: int) -> list[tuple[str, float]]:
        """
        Semantic search via ChromaDB, constrained to the current sandbox.

        Previously this fetched k*4 candidates from the GLOBAL collection
        and filtered down to the sandbox afterward with a Python loop —
        which meant a small/sparse sandbox could legitimately get back
        fewer than k results even when k relevant chunks existed, because
        the top global-ranked hits were dominated by other, unrelated
        sandboxes. That's the wrong order of operations.

        This version filters server-side via Chroma's `where` clause
        (`abs_path $in [...]`), so the similarity search itself only
        ever considers chunks belonging to this sandbox — filter first,
        then rank — rather than rank globally and filter after the fact.

        For large sandboxes, the file list is split into batches of
        CHROMA_QUERY_BATCH_SIZE (same reasoning as _fetch_chunks_for_paths:
        Chroma has no native path-prefix filter, only exact-match/$in, so
        a single sandbox-wide $in list could otherwise grow unbounded).
        Each batch returns its own top-k; results are merged and re-sorted
        by score afterward, since "top k overall" isn't the same as
        "top k from batch 1" when a sandbox spans multiple batches.

        Returns [(chunk_id, relevance_score), ...].
        Relevance score is normalised to [0, 1] by LangChain.
        """
        if not self._sandbox_abs_paths:
            return []

        paths = list(self._sandbox_abs_paths)
        all_hits: list[tuple] = []  # [(doc, score), ...] across all batches

        try:
            for i in range(0, len(paths), CHROMA_QUERY_BATCH_SIZE):
                batch = paths[i : i + CHROMA_QUERY_BATCH_SIZE]
                hits = self._vectorstore.similarity_search_with_relevance_scores(
                    query,
                    k=k,
                    filter={"abs_path": {"$in": batch}},
                )
                all_hits.extend(hits)
        except Exception as exc:
            log.warning("rag.semantic_search_failed", error=str(exc))
            return []

        # Merge across batches: highest relevance score first, then take
        # the overall top k (a single batch's top-k isn't necessarily the
        # sandbox's top-k once there's more than one batch).
        all_hits.sort(key=lambda pair: pair[1], reverse=True)

        results = []
        for doc, score in all_hits[:k]:
            meta     = doc.metadata
            abs_path = meta.get("abs_path", "")
            # Defensive second check — should always pass given the
            # `where` filter above, kept in case of Chroma version
            # quirks around filter semantics.
            if abs_path not in self._sandbox_abs_paths:
                continue
            chunk_id = self._chunk_id(abs_path, meta["chunk_index"])
            results.append((chunk_id, float(score)))
        return results


    @staticmethod
    def _rrf_merge(
        tfidf_hits:    list[tuple[str, float]],
        semantic_hits: list[tuple[str, float]],
        k: int = RRF_K,
    ) -> list[str]:
        """
        Reciprocal Rank Fusion.

        Each list contributes  1 / (k + rank)  to the shared score.
        A chunk appearing in both lists is rewarded twice.
        Returns chunk IDs sorted by descending RRF score.
        """
        rrf_scores: dict[str, float] = {}

        for rank, (chunk_id, _) in enumerate(tfidf_hits):
            rrf_scores[chunk_id] = rrf_scores.get(chunk_id, 0.0) + 1.0 / (k + rank + 1)

        for rank, (chunk_id, _) in enumerate(semantic_hits):
            rrf_scores[chunk_id] = rrf_scores.get(chunk_id, 0.0) + 1.0 / (k + rank + 1)

        return sorted(rrf_scores, key=lambda cid: rrf_scores[cid], reverse=True)


    def search(self, query: str, top_k: int = TOP_K) -> list[dict]:
        """
        Hybrid search: TF-IDF keyword pass  +  semantic pass  →  RRF merge.
        Both passes are pre-filtered to the current sandbox, so results
        are always scoped to the active session directory.

        Parameters
        ----------
        query   Natural-language user query.
        top_k   Max number of chunks to return.

        Returns
        -------
        List of result dicts, best match first:
        {
            "file_path":    "reports/Q3.pdf",   ← relative to current sandbox
            "abs_path":     "/home/.../Q3.pdf", ← absolute, for tool calls
            "file_name":    "Q3.pdf",
            "chunk_index":  3,
            "total_chunks": 12,
            "last_modified":"2024-01-15T10:30:00",
            "start_line":   84,
            "end_line":     99,
            "text":         "...the actual chunk content...",
        }
        """
        # Stage 1 — run both searches (both already sandbox-filtered)
        tfidf_hits    = self._tfidf_search(query,    k=TFIDF_CANDIDATES)
        semantic_hits = self._semantic_search(query, k=SEMANTIC_CANDIDATES)

        # Stage 2 — merge
        merged_ids = self._rrf_merge(tfidf_hits, semantic_hits)
        top_ids    = merged_ids[:top_k]

        if not top_ids:
            return []

        # Stage 3 — fetch full chunk data from ChromaDB
        raw = self._vectorstore.get(
            ids=top_ids,
            include=["documents", "metadatas"],
        )

        id_to_data = {
            cid: {"text": doc, "meta": meta}
            for cid, doc, meta in zip(
                raw["ids"], raw["documents"], raw["metadatas"]
            )
        }

        # Preserve RRF order; compute relative path at query time
        results = []
        for cid in top_ids:
            if cid not in id_to_data:
                continue
            entry    = id_to_data[cid]
            m        = entry["meta"]
            abs_path = m["abs_path"]

            # Derive display path relative to current sandbox
            try:
                rel_path = str(Path(abs_path).relative_to(self.sandbox_root))
            except ValueError:
                rel_path = abs_path  # fallback: shouldn't happen after sandbox filter

            results.append(
                {
                    "file_path":    rel_path,
                    "abs_path":     abs_path,
                    "file_name":    m["file_name"],
                    "chunk_index":  m["chunk_index"],
                    "total_chunks": m["total_chunks"],
                    "last_modified": m["last_modified"],
                    "start_line":   m.get("start_line"),
                    "end_line":     m.get("end_line"),
                    "text":         entry["text"],
                }
            )

        return results


    def format_results(self, results: list[dict]) -> str:
        """
        Format search results as a structured string for the LLM.
        Now includes precise line numbers so the LLM can call read_file_lines
        with exact ranges.
        """
        if not results:
            return "No relevant content found in the indexed files."

        lines = [f"Found {len(results)} relevant snippet(s):\n"]
        for i, r in enumerate(results, 1):
            line_info = ""
            if r.get("start_line") and r.get("end_line"):
                line_info = f"lines {r['start_line']}–{r['end_line']}  "

            lines.append(
                f"[{i}] {r['file_path']}  "
                f"{line_info}"
                f"(chunk {r['chunk_index'] + 1}/{r['total_chunks']}, "
                f"modified {r['last_modified']})\n"
                f"{r['text']}\n"
            )
        return "\n".join(lines)